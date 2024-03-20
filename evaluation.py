import re
import json
import os
import nltk
import existed_report
from nltk.tokenize import sent_tokenize
import pandas as pd
from model import category_specific
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

attributes = {'length', 'occasion', 'occassion', 'color', 'embellishment', 'neckline', 'fit', 'silhouette', 'style',
              'sleeve', 'fabric', 'detail', 'print and pattern'}


def get_attribute():
    attribute = set()
    path = 'data/all_brand_data_2019_2023/'
    directories = os.listdir(path)
    directories.sort()
    directories = directories[1:]
    for directory in directories:
        with open(path + directory + '/category_attribute_count.json', 'r') as json_file:
            json_data = json_file.read()
        data = dict(json.loads(json_data))
        for item in data.values():
            attribute = attribute.union(item.keys())
    with open('data/all_attribute.txt', 'w') as file:
        file.write(str(attribute))


def contains_number_except_year(sentence):
    # Regular expression to match any number except years in the range 1000-2999
    number_pattern = r'\b(?!(?:1000|1[0-9]{3}|2[0-9]{3})\b)\d+\b'
    return re.search(number_pattern, sentence) is not None


def contains_attributes(sentence):
    pattern = re.compile(r'\b(?:' + '|'.join([f'{keyword}s?' for keyword in attributes]) + r')\b', re.IGNORECASE)
    return re.search(pattern, sentence) is not None


def contain_keywords(sentence):
    keywords = {'decline', 'increase', 'rise', 'fall', 'decrease', 'growth'}
    pattern = re.compile(r'\b(?:' + '|'.join(keywords) + r')\b', re.IGNORECASE)
    return re.search(pattern, sentence) is not None


def extract_attribute(sentence):
    '''
    return the first str that contains any attribute
    :param sentence:
    :param attributes:
    :return:
    '''
    pattern = re.compile(r'\b(?:' + '|'.join([f'{keyword}s?' for keyword in attributes]) + r')\b', re.IGNORECASE)
    return re.search(pattern, sentence).group()


def label_statement(statement, category, count: dict):
    statement_class = list()
    for state in list(statement):
        if len(state) < 5:
            continue
        statement_with_class = dict()
        statement_with_class['statement'] = state
        if contains_number_except_year(state):
            statement_with_class['class1'] = 'number'
            print(state)
            count['number'] = count.get('number', 0) + 1
        else:
            statement_with_class['class1'] = 'non-number'
            count['non-number'] = count.get('non-number', 0) + 1
        if contains_attributes(state):
            statement_with_class['class2'] = extract_attribute(state)
        else:
            statement_with_class['class2'] = category
        count[statement_with_class['class2']] = count.get(statement_with_class['class2'], 0) + 1
        statement_class.append(statement_with_class)
    return statement_class, count


def statement_identification(description: list):
    statement = list()
    for sentence in description:
        if contains_number_except_year(sentence) or contains_attributes(sentence) or contain_keywords(
                sentence):
            statement.append(sentence)

    return statement


def statement_extraction(years, seasons, categories, brands, generative_model):
    statement_data = list()

    count = dict()
    with open('all_statement_v1.txt', 'w') as statement_file:
        for year in years:
            for season in seasons:
                for category in categories:
                    for b in brands:
                        brand = list()
                        brand.append(b)
                        if existed_report.check_exist_repost(year, season, category, brand, generative_model):
                            brand.sort()
                            if season == 'Spring/Summer (S/S)':
                                season = 'springsummer'
                            data = existed_report.load_file(year, season, category, brand, generative_model)
                            overview_description = data['description']
                            section_description = data['section_dict']['description']
                            # split text into sentences list
                            overview_description_sentences = sent_tokenize(overview_description)
                            overview_statement = statement_identification(overview_description_sentences)
                            overview_statement, count = label_statement(overview_statement, category, count)
                            # overview = {'description': overview_description, 'statement': overview_statement}
                            overview = {'statement': overview_statement}

                            section = list()
                            for section_item in section_description:
                                description = sent_tokenize(section_item['description'])
                                statement = statement_identification(description)
                                statement, count = label_statement(statement, category, count)
                                section_item['statement'] = statement
                                section.append(section_item)
                            statement_data.append(
                                {'name': '_'.join([year, season, category, str(brand), generative_model]),
                                 'year': year, 'season': season, 'category': category, 'brand': brand,
                                 'generative_model': generative_model, 'overview': overview, 'section': section})
        statement_file.write(str(statement_data) + str(count))
        return statement_data


def get_pics(year, category, brand, overview: bool):
    pics = list()
    if overview:
        path = 'data/charts/SS/line/'
        for item in os.listdir(path):
            if item == ('_'.join(['all_category', year, 'springsummer', str(brand)]) + '.png'):
                pics.append(path + item)
    if overview:
        sub_cate = category_specific[category]
        bar_pattern = re.compile(year + r'_(' + '|'.join(sub_cate) + r')_' + re.escape(str(brand)))
        pie_pattern = re.compile(r'(' + str(int(year) - 1) + '|' + year + r')' + r'_(' + '|'.join(
            sub_cate) + '|' + category + r')_' + re.escape(str(brand)))
    else:
        bar_pattern = re.compile(year + r'_(' + category + r')_' + re.escape(str(brand)))
        pie_pattern = re.compile(
            r'(' + str(int(year) - 1) + '|' + year + r')_' + category + '_' + re.escape(str(brand)))
    path = 'data/charts/SS/bar/'
    for item in os.listdir(path):
        if re.search(bar_pattern, item) is not None:
            pics.append(path + item)
    path = 'data/charts/SS/pie/'
    for item in os.listdir(path):
        if re.search(pie_pattern, item) is not None:
            pics.append(path + item)
    return pics


def add_img_to_sheet(sheet, imgs):
    previous_row_right = 1
    previous_row_left = sheet.max_row + 1
    previous_column_right = 'M'
    previous_column_left = 'C'
    for ind, img_path in enumerate(imgs):
        img = Image(img_path)
        img.height = 540
        img.width = 400
        if 'pie' in img_path:
            img.height = 350
        if previous_row_right > previous_row_left:
            col = 'A' if previous_column_left == 'C' else 'C'
            row = previous_row_left
            previous_column_left = 'A' if previous_column_left == 'C' else 'C'
            previous_row_left = row + 26 if previous_column_left == 'C' else previous_row_left
        else:
            col = 'G' if previous_column_right == 'M' else 'M'
            row = previous_row_right
            previous_column_right = 'G' if previous_column_right == 'M' else 'M'
            previous_row_right = row + 26 if previous_column_right == 'M' else previous_row_right
        sheet.add_image(img, col + str(row))


def save_as_excel():
    years = ['2020', '2021', '2022', '2023']
    seasons = ['Spring/Summer (S/S)']
    categories = ['Dresses&Skirts', 'Jackets&Coats&Outerwear', 'Topweights', 'Trousers&Shorts']
    brands = ['chanel', 'christian-dior', 'givenchy', 'louis-vuitton', 'saint-laurent', 'valentino']
    generative_model = 'GPT'
    datas = statement_extraction(years, seasons, categories, brands, generative_model)
    for i in range(0, len(datas)):
        data = datas[i]
        year = data['year']
        brand = data['brand']
        category = data['category']
        overview_sheet = pd.DataFrame(data['overview']['statement'])
        overview_sheet.drop(['class1', 'class2'], axis=1, inplace=True)
        overview_sheet['correctness score'] = None
        overview_sheet['significance score'] = None
        overview_sheet.to_excel("data/statement/" + data['name'] + '.xlsx', sheet_name="overview")
        writer = pd.ExcelWriter("data/statement/" + data['name'] + '.xlsx', mode="a", engine="openpyxl")
        for section in data['section']:
            section_sheet = pd.DataFrame(section['statement'])
            section_sheet.drop(['class1', 'class2'], axis=1, inplace=True)
            section_sheet['correctness score'] = None
            section_sheet['significance score'] = None
            section_sheet.to_excel(writer, sheet_name=section['section'])
        writer.close()

        # Insert image in the overview sheet
        workbook = load_workbook("data/statement/" + data['name'] + '.xlsx')
        overview_sheet = workbook["overview"]
        overview_sheet.column_dimensions['B'].width = 70
        overview_sheet.column_dimensions['C'].width = 18
        overview_sheet.column_dimensions['D'].width = 18
        imgs = get_pics(year, category, brand, True)
        for row in overview_sheet.iter_rows():
            for cell in row:
                cell.font = Font(size=15)
                cell.alignment = Alignment(wrap_text=True)
        add_img_to_sheet(overview_sheet, imgs)

        # Insert image in the section sheet
        for section in data['section']:
            section_sheet = workbook[section['section']]
            section_sheet.column_dimensions['B'].width = 70
            section_sheet.column_dimensions['C'].width = 18
            section_sheet.column_dimensions['D'].width = 18
            for row in section_sheet.iter_rows():
                for cell in row:
                    cell.font = Font(size=15)
                    cell.alignment = Alignment(wrap_text=True)
            imgs = get_pics(year, section['section'], brand, False)
            add_img_to_sheet(section_sheet, imgs)

        # Save the workbook
        workbook.save("data/statement_"+existed_report.prompt_version+"/" + data['name'] + '.xlsx')


# nltk.download('punkt')
# get_attribute()
save_as_excel()
